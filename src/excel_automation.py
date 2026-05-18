"""
excel_automation.py - Fills the BIR Form 2307 Excel template with payee data.
Preserves all formatting, merged cells, formulas and borders.
"""
import importlib
import logging
import shutil
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.config_loader import AppConfig
from src.validators import PayeeRecord

logger = logging.getLogger(__name__)


class ExcelAutomationError(Exception):
    """Raised for Excel fill/save failures."""


class BIR2307ExcelFiller:
    """
    Fills the BIR 2307 Excel template for a given PayeeRecord.
    Opens a fresh copy of the template for each payee.
    """

    def __init__(self, config: AppConfig):
        self._config = config
        self._cell_map = config.cell_map
        self._template = config.template_path
        self._use_com = self._check_com_available()
        logger.info(
            "Excel filler initialized using %s.",
            "Excel COM" if self._use_com else "openpyxl fallback"
        )

        if not self._template.exists():
            raise ExcelAutomationError(
                f"Template not found: {self._template}"
            )

    # ── Public API ────────────────────────────────────────────────────────

    def fill_and_save(self, record: PayeeRecord) -> Path:
        """
        Fill the BIR 2307 template with data from ``record`` and save
        to the appropriate output folder.

        Returns the saved .xlsx file path.
        """
        output_path = self._resolve_output_path(record)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Copy template to preserve the original
        shutil.copy2(self._template, output_path)

        sheet_name = self._cell_map.get("template_sheet", "Page1")
        if self._use_com:
            self._fill_with_com(output_path, sheet_name, record)
        else:
            wb = load_workbook(output_path)
            if sheet_name not in wb.sheetnames:
                raise ExcelAutomationError(
                    f"Sheet '{sheet_name}' not found in template. "
                    f"Available: {wb.sheetnames}"
                )
            ws: Worksheet = wb[sheet_name]
            self._fill_period(ws, record)
            self._fill_payee(ws, record)
            self._fill_payor(ws)
            self._fill_month_labels(ws, record)
            self._fill_income_details(ws, record)
            self._fill_signatory(ws)
            wb.save(output_path)

        logger.info("[%s] Excel saved → %s", record.payee_name, output_path)
        return output_path

    # ── Cell-fill helpers ─────────────────────────────────────────────────

    def _read(self, ws, cell_addr: str) -> str:
        """
        Read the current value of a cell.
        Supports both openpyxl worksheets and Excel COM worksheets (ws.Range).
        For openpyxl, resolves merged-cell top-left automatically.
        Returns a string (empty string if blank or on error).
        """
        if not cell_addr:
            return ""
        try:
            # ── Excel COM path ────────────────────────────────────────────
            if hasattr(ws, "Range"):
                val = ws.Range(cell_addr).Value
                return str(val) if val is not None else ""
            # ── openpyxl path ─────────────────────────────────────────────
            cell = ws[cell_addr]
            for merged in ws.merged_cells.ranges:
                if cell.coordinate in merged:
                    cell = ws.cell(row=merged.min_row, column=merged.min_col)
                    break
            return str(cell.value) if cell.value is not None else ""
        except Exception as exc:
            logger.warning("Could not read %s: %s", cell_addr, exc)
            return ""

    def _write(self, ws, cell_addr: str, value) -> None:
        """
        Write a value to a cell for either openpyxl or Excel COM.
        """
        if not cell_addr or value is None:
            return
        try:
            if hasattr(ws, "Range"):
                ws.Range(cell_addr).Value = value
                return
            cell = ws[cell_addr]
            # If this cell is part of a merged range, write to its top-left
            for merged in ws.merged_cells.ranges:
                if cell.coordinate in merged:
                    tl = ws.cell(row=merged.min_row, column=merged.min_col)
                    tl.value = value
                    return
            cell.value = value
        except Exception as exc:
            logger.warning("Could not write to %s: %s", cell_addr, exc)

    def _set_font_size(self, ws, cell_addr: str, size: int) -> None:
        """Set the font size for a target cell or merged range."""
        if not cell_addr or size is None:
            return
        try:
            if hasattr(ws, "Range"):
                rng = ws.Range(cell_addr)
                rng.Font.Size = size
                return
            cell = ws[cell_addr]
            for merged in ws.merged_cells.ranges:
                if cell.coordinate in merged:
                    cell = ws.cell(row=merged.min_row, column=merged.min_col)
                    break
            cell.font = Font(size=size)
        except Exception as exc:
            logger.warning("Could not set font size for %s: %s", cell_addr, exc)

    def _fill_period(self, ws: Worksheet, record: PayeeRecord) -> None:
        period = self._cell_map["period"]
        self._write(ws, period["date_from"], record.date_from)
        self._set_font_size(ws, period["date_from"], 12)
        self._write(ws, period["date_to"], record.date_to)
        self._set_font_size(ws, period["date_to"], 12)
        logger.debug(
            "Period: %s → %s", record.date_from, record.date_to
        )

    def _fill_payee(self, ws: Worksheet, record: PayeeRecord) -> None:
        payee_map = self._cell_map["payee"]
        
        # Split TIN into segments (XXX-XXX-XXX-XXX)
        tin_parts = record.payee_tin.split('-') if record.payee_tin else ['', '', '', '']
        tin_parts = (tin_parts + ['', '', '', ''])[:4]  # Pad with empty if needed
        
        self._write(ws, payee_map.get("tin1"), tin_parts[0])
        self._set_font_size(ws, payee_map.get("tin1"), 12)
        self._write(ws, payee_map.get("tin2"), tin_parts[1])
        self._set_font_size(ws, payee_map.get("tin2"), 12)
        self._write(ws, payee_map.get("tin3"), tin_parts[2])
        self._set_font_size(ws, payee_map.get("tin3"), 12)
        self._write(ws, payee_map.get("tin4"), tin_parts[3])
        self._set_font_size(ws, payee_map.get("tin4"), 12)
        
        self._write(ws, payee_map.get("name"), record.payee_name)
        self._set_font_size(ws, payee_map.get("name"), 12)
        # Address and zip are optional and left blank unless configured on the record.
        logger.debug("Payee info written: %s / %s", record.payee_name, record.payee_tin)

    def _fill_payor(self, ws) -> None:
        payor_cfg = self._config.payor
        payor_map = self._cell_map["payor"]
        
        # Split TIN into segments (XXX-XXX-XXX-XXX)
        tin_value = payor_cfg.get("tin", "")
        tin_parts = tin_value.split('-') if tin_value else ['', '', '', '']
        tin_parts = (tin_parts + ['', '', '', ''])[:4]  # Pad with empty if needed
        
        self._write(ws, payor_map.get("tin1"), tin_parts[0])
        self._set_font_size(ws, payor_map.get("tin1"), 12)
        self._write(ws, payor_map.get("tin2"), tin_parts[1])
        self._set_font_size(ws, payor_map.get("tin2"), 12)
        self._write(ws, payor_map.get("tin3"), tin_parts[2])
        self._set_font_size(ws, payor_map.get("tin3"), 12)
        self._write(ws, payor_map.get("tin4"), tin_parts[3])
        self._set_font_size(ws, payor_map.get("tin4"), 12)
        self._write(ws, payor_map.get("name"), payor_cfg.get("name"))
        self._write(ws, payor_map.get("address"), payor_cfg.get("address"))
        self._write(ws, payor_map.get("zip_code"), payor_cfg.get("zip_code"))
        logger.debug("Payor info written.")
        logger.debug("Payor info written.")

    def _fill_signatory(self, ws) -> None:
        # Signatory config: "Name / Designation / TIN"
        raw = self._config.payor.get("signatory", "")
        parts = [p.strip() for p in raw.split("/")]
        signatory_name = parts[0] if len(parts) > 0 else ""
        signatory_position = parts[1] if len(parts) > 1 else ""
        signatory_tin = parts[2] if len(parts) > 2 else ""

        signatory_map = self._cell_map.get("signatory", {})

        # ── name cell: "existing text / Name / Designation / TIN" ─────────
        # Read the existing template text, then append the three parts
        # separated by "/" so nothing is overwritten.
        name_addr = signatory_map.get("name")
        existing_name = self._read(ws, name_addr)
        new_parts = [p for p in [signatory_name, signatory_position, signatory_tin] if p]
        combined_name = (
            f"{existing_name} / {' / '.join(new_parts)}".strip(" /")
            if existing_name
            else " / ".join(new_parts)
        )
        self._write(ws, name_addr, combined_name)
        self._set_font_size(ws, name_addr, 12)

        logger.debug(
            "Signatory written (concatenated) — '%s'",
            combined_name,
        )

    def _fill_month_labels(self, ws, record: PayeeRecord) -> None:
        labels = self._cell_map.get("month_labels", {})
        cols = labels.get("columns", {})
        self._write(ws, f"{cols.get('month1_label')}{labels.get('row')}", record.month1_label)
        self._write(ws, f"{cols.get('month2_label')}{labels.get('row')}", record.month2_label)
        self._write(ws, f"{cols.get('month3_label')}{labels.get('row')}", record.month3_label)
        logger.debug(
            "Month labels written: %s, %s, %s",
            record.month1_label, record.month2_label, record.month3_label,
        )

    def _fill_income_details(self, ws, record: PayeeRecord) -> None:
        """Fill Part III income rows with ATC, amounts."""
        rows_cfg = self._cell_map["income_rows"]
        row_range = rows_cfg["row_range"]
        cols = rows_cfg["columns"]
        first_row = row_range[0]

        # Use the first available data row
        row = first_row

        def col_cell(col_letter: str) -> str:
            return f"{col_letter}{row}"

        self._write(ws, col_cell(cols["description"]),
                    record.income_description)
        self._write(ws, col_cell(cols["atc_code"]),
                    record.atc_code)
        self._write(ws, col_cell(cols["month1_amount"]),
                    record.month1_amount)
        self._write(ws, col_cell(cols["month2_amount"]),
                    record.month2_amount)
        self._write(ws, col_cell(cols["month3_amount"]),
                    record.month3_amount)

        total = (record.month1_amount
                 + record.month2_amount
                 + record.month3_amount)
        self._write(ws, col_cell(cols["total_amount"]), total)

        if record.tax_withheld is not None:
            self._write(ws, col_cell(cols["tax_withheld"]),
                        record.tax_withheld)
        else:
            computed_tax = total * record.tax_rate
            self._write(ws, col_cell(cols["tax_withheld"]), computed_tax)

        logger.debug(
            "Income row %d filled — ATC=%s M1=%.2f M2=%.2f M3=%.2f",
            row, record.atc_code,
            record.month1_amount, record.month2_amount, record.month3_amount,
        )

    # ── Output path ───────────────────────────────────────────────────────

    def _fill_with_com(
        self,
        output_path: Path,
        sheet_name: str,
        record: PayeeRecord,
    ) -> None:
        try:
            win32com = importlib.import_module("win32com.client")
        except ImportError as exc:
            raise ExcelAutomationError(
                "Excel COM not available for template fill."
            ) from exc

        excel = None
        wb = None
        try:
            excel = win32com.Dispatch("Excel.Application")
            try:
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.ScreenUpdating = False
            except Exception as e:
                logger.warning("Could not set Excel properties (may be running in restricted mode): %s", e)

            wb = excel.Workbooks.Open(str(output_path.resolve()))
            ws = wb.Worksheets(sheet_name)

            self._fill_period(ws, record)
            self._fill_payee(ws, record)
            self._fill_payor(ws)
            self._fill_month_labels(ws, record)
            self._fill_income_details(ws, record)
            self._fill_signatory(ws)

            wb.Save()
        except Exception as exc:
            raise ExcelAutomationError(
                f"Excel COM fill failed: {exc}"
            ) from exc
        finally:
            if wb:
                try:
                    wb.Close(SaveChanges=True)
                except Exception:
                    pass
            if excel:
                try:
                    excel.Quit()
                except Exception:
                    pass

    def _check_com_available(self) -> bool:
        """
        Return True when win32com is importable (Windows + pywin32 installed).
        COM MUST be used when the template contains images, logos, or shapes
        because openpyxl strips them on save.
        Falls back to openpyxl only when COM is genuinely unavailable.
        """
        try:
            importlib.import_module("win32com.client")
            logger.info("win32com available — using Excel COM to preserve images/logos.")
            return True
        except ImportError:
            logger.warning(
                "win32com not available. Falling back to openpyxl — "
                "WARNING: embedded images and logos will be lost from the output."
            )
            return False

    def _resolve_output_path(self, record: PayeeRecord) -> Path:
        """
        Build output path:
          OUTPUT / Q1_2026 / Juan Dela Cruz.xlsx
        """
        folder_name = f"{record.quarter}_{record.year}"
        safe_name = _sanitize_filename(record.payee_name)
        base = self._config.output_base / folder_name
        return base / f"{safe_name}.xlsx"


def _sanitize_filename(name: str) -> str:
    """Strip characters that are illegal in Windows filenames."""
    illegal = r'\/:*?"<>|'
    for ch in illegal:
        name = name.replace(ch, "_")
    return name.strip().strip(".")